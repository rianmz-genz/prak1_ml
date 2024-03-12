from openpyxl import Workbook
import csv

def header_column(sheet, column):
    for col_idx, header in enumerate(column, start=1):
        sheet.cell(row=1, column=col_idx).value = header

def write_data(sheet, data):
    for row_idx, row in enumerate(data, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx).value = value
            
def data_to_excel(df, column, data, filepath):
    # Print the DataFrame
    print(df)

    # Create a new Workbook object
    wb = Workbook()
    # Get the active worksheet
    sheet = wb.active

    # Write column mahasiswa to the first row
    header_column(sheet, column)
    # Write data to the worksheet starting from the second row
    write_data(sheet, data)

    # Save the workbook
    wb.save(filepath)

def read_csv(filepath, limit = 10):
    count = 0
    # Mengimpor data dari file CSV

    # Membuka file CSV
    with open(filepath, newline='') as csvfile:
        # Membaca file CSV
        csv_reader = csv.reader(csvfile)
        
        # Membaca dan menampilkan header
        header = next(csv_reader)
        print("Header:", header)
        
        # Membaca dan menampilkan setiap baris data
        for row in csv_reader:
            print(row)
            count += 1
            if count >= limit:
                break

